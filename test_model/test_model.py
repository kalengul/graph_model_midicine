"""Программа для тестирования языковых моделей"""
from sklearn.model_selection import train_test_split
from transformers import AutoModelForTokenClassification, AutoTokenizer
from datasets import load_metric
import torch
from openpyxl import Workbook
import json
import pandas as pd


# Загрузка данных из JSON
def load_json_data(filename):
    with open(filename, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data

# Преобразование данных в pandas DataFrame
def convert_to_dataframe(data):
    rows = []
    for entry in data:
        tokens = entry['tokens']
        tags = entry['tags']
        rows.append({'tokens': tokens, 'tags': tags})
    return pd.DataFrame(rows)

# функция сшивания  токенов BERT разорваных слов 
def merge_BERT_tokens(tokens, labels):
    """
    Функция для объединения подслов в исходные слова и сохранения меток.

    tokens: Список токенов (токенизированные слова).
    labels: Список меток для каждого токена.

    Возвращает список слов с соответствующими метками.
    """
    merged_tokens = []
    merged_labels = []
    
    current_token = ""
    current_label = labels[0]
    
    for token, label in zip(tokens, labels):
        if token.startswith("##"):
            current_token += token[2:]  # Убираем префикс '##' и добавляем к текущему слову
        else:
            if current_token:  # Если есть собранное слово, добавляем его в результат
                merged_tokens.append(current_token)
                merged_labels.append(current_label)
            current_token = token  # Начинаем новое слово
            current_label = label
    
    # Добавляем последний токен
    if current_token:
        merged_tokens.append(current_token)
        merged_labels.append(current_label)
    
    return merged_tokens, merged_labels


def merge_T5_tokens(tokens, labels):
    """
    Функция для объединения подслов в исходные слова и сохранения меток.

    tokens: Список токенов (токенизированные слова).
    labels: Список меток для каждого токена.

    Возвращает список слов с соответствующими метками.
    """
    merged_tokens = []
    merged_labels = []
    
    current_token = ""
    current_label = labels[0]
    
    for token, label in zip(tokens, labels):
        if token.startswith("▁"):  # T5 использует символ "▁" для обозначения начала нового слова
            if current_token:  # Если есть собранное слово, добавляем его в результат
                merged_tokens.append(current_token)
                merged_labels.append(current_label)
            current_token = token[1:]  # Убираем символ "▁" и начинаем новое слово
            current_label = label
        else:
            current_token += token  # Добавляем подслово к текущему слову
    
    # Добавляем последний токен
    if current_token:
        merged_tokens.append(current_token)
        merged_labels.append(current_label)
    
    return merged_tokens, merged_labels


wb = Workbook()
# делаем единственный лист активным 
ws = wb.active

texts = ['Эффективность арипипразола при адъюнктивном лечении БДР была продемонстрирована в двух \
        краткосрочных плацебо-контролируемых исследованиях у взрослых пациентов, соответствующих критериям DSM-IV для БДР, \
        у которых был неадекватный ответ на предыдущее лечение антидепрессантами в текущем эпизоде и которые также \
        продемонстрировали неадекватный ответ на 8-недельную проспективную терапию антидепрессантами.',
        '''
        Фармакология
        Ингибирует синтез ПГ и снижает возбудимость центра терморегуляции гипоталамуса. Быстро всасывается из ЖКТ, связывается с белками плазмы. \
        T1/2 из плазмы 1–4 ч. Метаболизируется в печени с образованием глюкуронида и сульфата парацетамола. Выводится почками главным образом \
        в виде продуктов конъюгации, менее 5% экскретируется в неизмененном виде.
        Применение вещества Парацетамол
        Боли слабой и умеренной интенсивности (головная и зубная боль, мигрень, боль в спине, артралгия, миалгия, невралгия, меналгия), \
        лихорадочный синдром при простудных заболеваниях.
        Противопоказания
        Гиперчувствительность, нарушение функций почек и печени, алкоголизм, детский возраст (до 6 лет).
        Побочные действия вещества Парацетамол
        Агранулоцитоз, тромбоцитопения, анемия, почечная колика, асептическая пиурия, интерстициальный гломерулонефрит, аллергические реакции в виде кожных высыпаний.
        Взаимодействие
        Увеличивает эффект непрямых антикоагулянтов (производных кумарина) и вероятность поражения печени гепатотоксичными препаратами. Метоклопрамид повышает, \
        а колестирамин снижает скорость всасывания. Барбитураты уменьшают жаропонижающую активность.
        Передозировка
        Симптомы: в первые 24 ч — бледность, тошнота, рвота и боль в абдоминальной области; через 12–48 ч — повреждения почек и печени \
        с развитием печеночной недостаточности (энцефалопатия, кома, летальный исход), сердечные аритмии и панкреатит. Поражение печени возможны \
        при приеме 10 г и более (у взрослых).
        Лечение: назначение метионина внутрь или в/в введение N-ацетилцистеина.
        ''',
        'Ингибирует синтез ПГ и снижает возбудимость центра терморегуляции гипоталамуса. Быстро всасывается из ЖКТ, связывается с белками плазмы. \
        T1/2 из плазмы 1–4 ч. Метаболизируется в печени с образованием глюкуронида и сульфата парацетамола. \
        Выводится почками главным образом в виде продуктов конъюгации, менее 5% экскретируется в неизмененном виде.',
        ]


# Пример меток
#label_list = ['O', 'B-noun', "I-noun", 'B-glag', 'I-glag']
#label2id = {label: i for i, label in enumerate(label_list)}
#id2label = {i: label for i, label in enumerate(label_list)}


BIO_mark_sent_v2_rubert_base_cased_model_names = ['saved_model_BIO_mark_sent_v2_rubert-base-cased_3', 
                                                  'saved_model_BIO_mark_sent_v2_rubert-base-cased_4',
                                                  'saved_model_BIO_mark_sent_v2_rubert-base-cased_5',
                                                  'saved_model_BIO_mark_sent_v2_rubert-base-cased_6',]


BIO_mark_sent_v2_RuBioBERT_model_names = ['saved_model_BIO_mark_sent_v2_RuBioBERT_3', 
                                          'saved_model_BIO_mark_sent_v2_RuBioBERT_4',
                                          'saved_model_BIO_mark_sent_v2_RuBioBERT_5',
                                          'saved_model_BIO_mark_sent_v2_RuBioBERT_6',]


BIO_mark_sent_v2_rut5_base_model_names = ['saved_model_BIO_mark_sent_v2_rut5-base_3',
                                          'saved_model_BIO_mark_sent_v2_rut5-base_4',
                                          'saved_model_BIO_mark_sent_v2_rut5-base_5',
                                          'saved_model_BIO_mark_sent_v2_rut5-base_6',]


BIO_mark_sent_v2_rut5_base_multitask_model_names = ['saved_model_BIO_mark_sent_v2_rut5-base-multitask_3',
                                                    'saved_model_BIO_mark_sent_v2_rut5-base-multitask_4',
                                                    'saved_model_BIO_mark_sent_v2_rut5-base-multitask_5',
                                                    'saved_model_BIO_mark_sent_v2_rut5-base-multitask_6',]


annotation = 'BIO_mark_sent_v2.json'

data = load_json_data(annotation)
df = convert_to_dataframe(data)

label_list = list()
for i in range(df.shape[0]):
    label_list.extend((list(df.tags.iloc[i])))

label_list = set(label_list)


first_row = 1
count = 0
for text in texts:

  count += 1

  column_token_rubert = 1
  column_ner_rubert = 2
  column_token_rubiobert = 3
  column_ner_rubiobert = 4

  column_token_rut5 = 5
  column_ner_rut5 = 6
  column_token_rut5_multitask = 7
  column_ner_rut5_multitask = 8
  
  
  for model_name1, model_name2, model_name3, model_name4 in zip(BIO_mark_sent_v2_rubert_base_cased_model_names, 
                                                                BIO_mark_sent_v2_RuBioBERT_model_names,
                                                                BIO_mark_sent_v2_rut5_base_model_names,
                                                                BIO_mark_sent_v2_rut5_base_multitask_model_names):
    number_row = 2
    # Загрузка токенизатора и модели
    tokenizer1 = AutoTokenizer.from_pretrained(model_name1, trust_remote_code=True)
    model1 = AutoModelForTokenClassification.from_pretrained(model_name1, num_labels=len(label_list), trust_remote_code=True)

    inputs1 = tokenizer1(text, return_tensors="pt")

    with torch.no_grad():
      logits = model1(**inputs1).logits

    predictions1 = torch.argmax(logits, dim=2)
    predicted_token_class1 = [model1.config.id2label[t.item()] for t in predictions1[0]]

    tokenizer2 = AutoTokenizer.from_pretrained(model_name2, trust_remote_code=True)
    model2 = AutoModelForTokenClassification.from_pretrained(model_name2, num_labels=len(label_list), trust_remote_code=True)

    inputs2 = tokenizer2(text, return_tensors="pt")

    with torch.no_grad():
      logits = model2(**inputs2).logits

    predictions2 = torch.argmax(logits, dim=2)
    predicted_token_class2 = [model2.config.id2label[t.item()] for t in predictions2[0]]
    

    # Загрузка токенизатора и модели
    tokenizer3 = AutoTokenizer.from_pretrained(model_name3, trust_remote_code=True)
    model3 = AutoModelForTokenClassification.from_pretrained(model_name3, num_labels=len(label_list), trust_remote_code=True)

    inputs3 = tokenizer3(text, return_tensors="pt")

    with torch.no_grad():
      logits = model3(**inputs3).logits

    predictions3 = torch.argmax(logits, dim=2)
    predicted_token_class3 = [model3.config.id2label[t.item()] for t in predictions3[0]]

    tokenizer4 = AutoTokenizer.from_pretrained(model_name4, trust_remote_code=True)
    model4 = AutoModelForTokenClassification.from_pretrained(model_name4, num_labels=len(label_list), trust_remote_code=True)

    inputs4 = tokenizer4(text, return_tensors="pt")

    with torch.no_grad():
      logits = model4(**inputs4).logits

    predictions4 = torch.argmax(logits, dim=2)
    predicted_token_class4 = [model4.config.id2label[t.item()] for t in predictions4[0]]
    

    ws.cell(row=first_row, column=column_token_rubert, value=model_name1)
    ws.cell(row=first_row, column=column_token_rubiobert, value=model_name2)

    ws.cell(row=first_row, column=column_token_rut5, value=model_name3)
    ws.cell(row=first_row, column=column_token_rut5_multitask, value=model_name4)

    print('column_token_rubert =', column_token_rubert)
    print('column_ner_rubert =', column_ner_rubert)

    print('column_token_rut5 =', column_token_rut5)
    print('column_ner_rut5 =', column_ner_rut5)

    print('column_token_rut5_multitask =', column_token_rut5_multitask)
    print('column_ner_rut5_multitask =', column_ner_rut5_multitask)

    ws.merge_cells(start_row=first_row, start_column=column_token_rubert, end_row=first_row, end_column=column_ner_rubert)
    ws.merge_cells(start_row=first_row, start_column=column_token_rubiobert, end_row=first_row, end_column=column_ner_rubiobert)

    tokens1 = tokenizer1.convert_ids_to_tokens(inputs1["input_ids"][0])
    tokens2 = tokenizer2.convert_ids_to_tokens(inputs2["input_ids"][0])

    ws.merge_cells(start_row=first_row, start_column=column_token_rut5, end_row=first_row, end_column=column_ner_rut5)
    ws.merge_cells(start_row=first_row, start_column=column_token_rut5_multitask, end_row=first_row, end_column=column_ner_rut5_multitask)

    tokens3 = tokenizer3.convert_ids_to_tokens(inputs3["input_ids"][0])
    tokens4 = tokenizer4.convert_ids_to_tokens(inputs4["input_ids"][0])

    for source_text1, pred1, source_text2, pred2, \
        source_text3, pred3, source_text4, pred4 in zip(tokens1, predicted_token_class1, 
                                                        tokens2, predicted_token_class2,
                                                        tokens3, predicted_token_class3,
                                                        tokens4, predicted_token_class4,
                                                        ):

      ws.cell(row=number_row, column=column_token_rubert, value=source_text1)
      ws.cell(row=number_row, column=column_ner_rubert, value=pred1)
      ws.cell(row=number_row, column=column_token_rubiobert, value=source_text2)
      ws.cell(row=number_row, column=column_ner_rubiobert, value=pred2)

      ws.cell(row=number_row, column=column_token_rut5, value=source_text3)
      ws.cell(row=number_row, column=column_ner_rut5, value=pred3)
      ws.cell(row=number_row, column=column_token_rut5_multitask, value=source_text4)
      ws.cell(row=number_row, column=column_ner_rut5_multitask, value=pred4)

      number_row += 1

    column_token_rubert += 8
    column_ner_rubert += 8
    column_token_rubiobert += 8
    column_ner_rubiobert += 8

    column_token_rut5 += 8
    column_ner_rut5 += 8
    column_token_rut5_multitask += 8
    column_ner_rut5_multitask += 8

  wb.save(f'comperison_BIO_NER_model_{count}.xlsx')




mark_sent_v2_rubert_base_cased_model_names = ['saved_model_mark_sent_2_rubert-base-cased_3', 
                                              'saved_model_mark_sent_2_rubert-base-cased_4',
                                              'saved_model_mark_sent_2_rubert-base-cased_5',
                                              'saved_model_mark_sent_2_rubert-base-cased_6',]


mark_sent_v2_RuBioBERT_model_names = ['saved_model_mark_sent_2_RuBioBERT_3', 
                                      'saved_model_mark_sent_2_RuBioBERT_4',
                                      'saved_model_mark_sent_2_RuBioBERT_5',
                                      'saved_model_mark_sent_2_RuBioBERT_6',]


mark_sent_v2_rut5_base_model_names = ['saved_model_mark_sent_2_rut5-base_3',
                                      'saved_model_mark_sent_2_rut5-base_4',
                                      'saved_model_mark_sent_2_rut5-base_5',
                                      'saved_model_mark_sent_2_rut5-base_6',]


mark_sent_v2_rut5_base_multitask_model_names = ['saved_model_mark_sent_2_rut5-base-multitask_3',
                                                'saved_model_mark_sent_2_rut5-base-multitask_4',
                                                'saved_model_mark_sent_2_rut5-base-multitask_5',
                                                'saved_model_mark_sent_2_rut5-base-multitask_6',]

wb = Workbook()
# делаем единственный лист активным 
ws = wb.active

annotation = 'mark_sent_2.json'

data = load_json_data(annotation)
df = convert_to_dataframe(data)

label_list = list()
for i in range(df.shape[0]):
    label_list.extend((list(df.tags.iloc[i])))

label_list = set(label_list)


first_row = 1
count = 0
for text in texts:

  count += 1

  column_token_rubert = 1
  column_ner_rubert = 2
  column_token_rubiobert = 3
  column_ner_rubiobert = 4

  column_token_rut5 = 5
  column_ner_rut5 = 6
  column_token_rut5_multitask = 7
  column_ner_rut5_multitask = 8
  
  
  for model_name1, model_name2, model_name3, model_name4 in zip(mark_sent_v2_rubert_base_cased_model_names, 
                                                                mark_sent_v2_RuBioBERT_model_names,
                                                                mark_sent_v2_rut5_base_model_names,
                                                                mark_sent_v2_rut5_base_multitask_model_names):
    number_row = 2
    # Загрузка токенизатора и модели
    tokenizer1 = AutoTokenizer.from_pretrained(model_name1, trust_remote_code=True)
    model1 = AutoModelForTokenClassification.from_pretrained(model_name1, num_labels=len(label_list), trust_remote_code=True)

    inputs1 = tokenizer1(text, return_tensors="pt")

    with torch.no_grad():
      logits = model1(**inputs1).logits

    predictions1 = torch.argmax(logits, dim=2)
    predicted_token_class1 = [model1.config.id2label[t.item()] for t in predictions1[0]]

    tokenizer2 = AutoTokenizer.from_pretrained(model_name2, trust_remote_code=True)
    model2 = AutoModelForTokenClassification.from_pretrained(model_name2, num_labels=len(label_list), trust_remote_code=True)

    inputs2 = tokenizer2(text, return_tensors="pt")

    with torch.no_grad():
      logits = model2(**inputs2).logits

    predictions2 = torch.argmax(logits, dim=2)
    predicted_token_class2 = [model2.config.id2label[t.item()] for t in predictions2[0]]
    

    # Загрузка токенизатора и модели
    tokenizer3 = AutoTokenizer.from_pretrained(model_name3, trust_remote_code=True)
    model3 = AutoModelForTokenClassification.from_pretrained(model_name3, num_labels=len(label_list), trust_remote_code=True)

    inputs3 = tokenizer3(text, return_tensors="pt")

    with torch.no_grad():
      logits = model3(**inputs3).logits

    predictions3 = torch.argmax(logits, dim=2)
    predicted_token_class3 = [model3.config.id2label[t.item()] for t in predictions3[0]]

    tokenizer4 = AutoTokenizer.from_pretrained(model_name4, trust_remote_code=True)
    model4 = AutoModelForTokenClassification.from_pretrained(model_name4, num_labels=len(label_list), trust_remote_code=True)

    inputs4 = tokenizer4(text, return_tensors="pt")

    with torch.no_grad():
      logits = model4(**inputs4).logits

    predictions4 = torch.argmax(logits, dim=2)
    predicted_token_class4 = [model4.config.id2label[t.item()] for t in predictions4[0]]
    

    ws.cell(row=first_row, column=column_token_rubert, value=model_name1)
    ws.cell(row=first_row, column=column_token_rubiobert, value=model_name2)

    ws.cell(row=first_row, column=column_token_rut5, value=model_name3)
    ws.cell(row=first_row, column=column_token_rut5_multitask, value=model_name4)

    print('column_token_rubert =', column_token_rubert)
    print('column_ner_rubert =', column_ner_rubert)

    print('column_token_rubiobert =', column_token_rubiobert)
    print('column_ner_rubiobert =', column_ner_rubiobert)

    print('column_token_rut5 =', column_token_rut5)
    print('column_ner_rut5 =', column_ner_rut5)

    print('column_token_rut5_multitask =', column_token_rut5_multitask)
    print('column_ner_rut5_multitask =', column_ner_rut5_multitask)

    ws.merge_cells(start_row=first_row, start_column=column_token_rubert, end_row=first_row, end_column=column_ner_rubert)
    ws.merge_cells(start_row=first_row, start_column=column_token_rubiobert, end_row=first_row, end_column=column_ner_rubiobert)

    tokens1 = tokenizer1.convert_ids_to_tokens(inputs1["input_ids"][0])
    tokens2 = tokenizer2.convert_ids_to_tokens(inputs2["input_ids"][0])

    ws.merge_cells(start_row=first_row, start_column=column_token_rut5, end_row=first_row, end_column=column_ner_rut5)
    ws.merge_cells(start_row=first_row, start_column=column_token_rut5_multitask, end_row=first_row, end_column=column_ner_rut5_multitask)

    tokens3 = tokenizer3.convert_ids_to_tokens(inputs3["input_ids"][0])
    tokens4 = tokenizer4.convert_ids_to_tokens(inputs4["input_ids"][0])

    for source_text1, pred1, source_text2, pred2, \
        source_text3, pred3, source_text4, pred4 in zip(tokens1, predicted_token_class1, 
                                                        tokens2, predicted_token_class2,
                                                        tokens3, predicted_token_class3,
                                                        tokens4, predicted_token_class4,
                                                        ):

      ws.cell(row=number_row, column=column_token_rubert, value=source_text1)
      ws.cell(row=number_row, column=column_ner_rubert, value=pred1)
      ws.cell(row=number_row, column=column_token_rubiobert, value=source_text2)
      ws.cell(row=number_row, column=column_ner_rubiobert, value=pred2)

      ws.cell(row=number_row, column=column_token_rut5, value=source_text3)
      ws.cell(row=number_row, column=column_ner_rut5, value=pred3)
      ws.cell(row=number_row, column=column_token_rut5_multitask, value=source_text4)
      ws.cell(row=number_row, column=column_ner_rut5_multitask, value=pred4)

      number_row += 1

    column_token_rubert += 8
    column_ner_rubert += 8
    column_token_rubiobert += 8
    column_ner_rubiobert += 8

    column_token_rut5 += 8
    column_ner_rut5 += 8
    column_token_rut5_multitask += 8
    column_ner_rut5_multitask += 8

  wb.save(f'comperison_NER_model_{count}.xlsx')


# работа BERT с последующим сшиванием
BIO_mark_sent_v2_rubert_base_cased_model_names = ['saved_model_BIO_mark_sent_v2_rubert-base-cased_3', 
                                                  'saved_model_BIO_mark_sent_v2_rubert-base-cased_4',
                                                  'saved_model_BIO_mark_sent_v2_rubert-base-cased_5',
                                                  'saved_model_BIO_mark_sent_v2_rubert-base-cased_6',]


BIO_mark_sent_v2_RuBioBERT_model_names = ['saved_model_BIO_mark_sent_v2_RuBioBERT_3', 
                                          'saved_model_BIO_mark_sent_v2_RuBioBERT_4',
                                          'saved_model_BIO_mark_sent_v2_RuBioBERT_5',
                                          'saved_model_BIO_mark_sent_v2_RuBioBERT_6',]


BIO_mark_sent_v2_rut5_base_model_names = ['saved_model_BIO_mark_sent_v2_rut5-base_3',
                                          'saved_model_BIO_mark_sent_v2_rut5-base_4',
                                          'saved_model_BIO_mark_sent_v2_rut5-base_5',
                                          'saved_model_BIO_mark_sent_v2_rut5-base_6',]


BIO_mark_sent_v2_rut5_base_multitask_model_names = ['saved_model_BIO_mark_sent_v2_rut5-base-multitask_3',
                                                    'saved_model_BIO_mark_sent_v2_rut5-base-multitask_4',
                                                    'saved_model_BIO_mark_sent_v2_rut5-base-multitask_5',
                                                    'saved_model_BIO_mark_sent_v2_rut5-base-multitask_6',]


annotation = 'BIO_mark_sent_v2.json'

data = load_json_data(annotation)
df = convert_to_dataframe(data)

label_list = list()
for i in range(df.shape[0]):
    label_list.extend((list(df.tags.iloc[i])))

label_list = set(label_list)


first_row = 1
count = 0
for text in texts:

  count += 1

  column_token_rubert = 1
  column_ner_rubert = 2
  column_token_rubiobert = 3
  column_ner_rubiobert = 4

  column_token_rut5 = 5
  column_ner_rut5 = 6
  column_token_rut5_multitask = 7
  column_ner_rut5_multitask = 8
  
  
  for model_name1, model_name2, model_name3, model_name4 in zip(BIO_mark_sent_v2_rubert_base_cased_model_names, 
                                                                BIO_mark_sent_v2_RuBioBERT_model_names,
                                                                BIO_mark_sent_v2_rut5_base_model_names,
                                                                BIO_mark_sent_v2_rut5_base_multitask_model_names):
    number_row = 2
    # Загрузка токенизатора и модели
    tokenizer1 = AutoTokenizer.from_pretrained(model_name1, trust_remote_code=True)
    model1 = AutoModelForTokenClassification.from_pretrained(model_name1, num_labels=len(label_list), trust_remote_code=True)

    inputs1 = tokenizer1(text, return_tensors="pt")

    with torch.no_grad():
      logits = model1(**inputs1).logits

    predictions1 = torch.argmax(logits, dim=2)
    predicted_token_class1 = [model1.config.id2label[t.item()] for t in predictions1[0]]

    tokenizer2 = AutoTokenizer.from_pretrained(model_name2, trust_remote_code=True)
    model2 = AutoModelForTokenClassification.from_pretrained(model_name2, num_labels=len(label_list), trust_remote_code=True)

    inputs2 = tokenizer2(text, return_tensors="pt")

    with torch.no_grad():
      logits = model2(**inputs2).logits

    predictions2 = torch.argmax(logits, dim=2)
    predicted_token_class2 = [model2.config.id2label[t.item()] for t in predictions2[0]]
    

    # Загрузка токенизатора и модели
    tokenizer3 = AutoTokenizer.from_pretrained(model_name3, trust_remote_code=True)
    model3 = AutoModelForTokenClassification.from_pretrained(model_name3, num_labels=len(label_list), trust_remote_code=True)

    inputs3 = tokenizer3(text, return_tensors="pt")

    with torch.no_grad():
      logits = model3(**inputs3).logits

    predictions3 = torch.argmax(logits, dim=2)
    predicted_token_class3 = [model3.config.id2label[t.item()] for t in predictions3[0]]

    tokenizer4 = AutoTokenizer.from_pretrained(model_name4, trust_remote_code=True)
    model4 = AutoModelForTokenClassification.from_pretrained(model_name4, num_labels=len(label_list), trust_remote_code=True)

    inputs4 = tokenizer4(text, return_tensors="pt")

    with torch.no_grad():
      logits = model4(**inputs4).logits

    predictions4 = torch.argmax(logits, dim=2)
    predicted_token_class4 = [model4.config.id2label[t.item()] for t in predictions4[0]]
    

    ws.cell(row=first_row, column=column_token_rubert, value=model_name1)
    ws.cell(row=first_row, column=column_token_rubiobert, value=model_name2)

    ws.cell(row=first_row, column=column_token_rut5, value=model_name3)
    ws.cell(row=first_row, column=column_token_rut5_multitask, value=model_name4)

    print('column_token_rubert =', column_token_rubert)
    print('column_ner_rubert =', column_ner_rubert)

    print('column_token_rut5 =', column_token_rut5)
    print('column_ner_rut5 =', column_ner_rut5)

    print('column_token_rut5_multitask =', column_token_rut5_multitask)
    print('column_ner_rut5_multitask =', column_ner_rut5_multitask)

    ws.merge_cells(start_row=first_row, start_column=column_token_rubert, end_row=first_row, end_column=column_ner_rubert)
    ws.merge_cells(start_row=first_row, start_column=column_token_rubiobert, end_row=first_row, end_column=column_ner_rubiobert)

    tokens1 = tokenizer1.convert_ids_to_tokens(inputs1["input_ids"][0])
    tokens2 = tokenizer2.convert_ids_to_tokens(inputs2["input_ids"][0])

    ws.merge_cells(start_row=first_row, start_column=column_token_rut5, end_row=first_row, end_column=column_ner_rut5)
    ws.merge_cells(start_row=first_row, start_column=column_token_rut5_multitask, end_row=first_row, end_column=column_ner_rut5_multitask)

    tokens3 = tokenizer3.convert_ids_to_tokens(inputs3["input_ids"][0])
    tokens4 = tokenizer4.convert_ids_to_tokens(inputs4["input_ids"][0])
    

    tokens1, predicted_token_class1 = merge_BERT_tokens(tokens1, predicted_token_class1)
    tokens2, predicted_token_class2 = merge_BERT_tokens(tokens2, predicted_token_class2)
    tokens3, predicted_token_class3 = merge_T5_tokens(tokens3, predicted_token_class3)
    tokens4, predicted_token_class4 = merge_T5_tokens(tokens4, predicted_token_class4)

    for source_text1, pred1, source_text2, pred2, \
        source_text3, pred3, source_text4, pred4 in zip(tokens1, predicted_token_class1, 
                                                        tokens2, predicted_token_class2,
                                                        tokens3, predicted_token_class3,
                                                        tokens4, predicted_token_class4,
                                                        ):

      ws.cell(row=number_row, column=column_token_rubert, value=source_text1)
      ws.cell(row=number_row, column=column_ner_rubert, value=pred1)
      ws.cell(row=number_row, column=column_token_rubiobert, value=source_text2)
      ws.cell(row=number_row, column=column_ner_rubiobert, value=pred2)

      ws.cell(row=number_row, column=column_token_rut5, value=source_text3)
      ws.cell(row=number_row, column=column_ner_rut5, value=pred3)
      ws.cell(row=number_row, column=column_token_rut5_multitask, value=source_text4)
      ws.cell(row=number_row, column=column_ner_rut5_multitask, value=pred4)

      number_row += 1

    column_token_rubert += 8
    column_ner_rubert += 8
    column_token_rubiobert += 8
    column_ner_rubiobert += 8

    column_token_rut5 += 8
    column_ner_rut5 += 8
    column_token_rut5_multitask += 8
    column_ner_rut5_multitask += 8

  wb.save(f'comperison_stitched_BIO_NER_model_{count}.xlsx')



# работа BERT с последующим сшиванием
BIO_mark_sent_v2_rubert_base_cased_model_names = ['saved_model_BIO_mark_sent_v2_rubert-base-cased_3', 
                                                  'saved_model_BIO_mark_sent_v2_rubert-base-cased_4',
                                                  'saved_model_BIO_mark_sent_v2_rubert-base-cased_5',
                                                  'saved_model_BIO_mark_sent_v2_rubert-base-cased_6',]


BIO_mark_sent_v2_RuBioBERT_model_names = ['saved_model_BIO_mark_sent_v2_RuBioBERT_3', 
                                          'saved_model_BIO_mark_sent_v2_RuBioBERT_4',
                                          'saved_model_BIO_mark_sent_v2_RuBioBERT_5',
                                          'saved_model_BIO_mark_sent_v2_RuBioBERT_6',]


BIO_mark_sent_v2_rut5_base_model_names = ['saved_model_BIO_mark_sent_v2_rut5-base_3',
                                          'saved_model_BIO_mark_sent_v2_rut5-base_4',
                                          'saved_model_BIO_mark_sent_v2_rut5-base_5',
                                          'saved_model_BIO_mark_sent_v2_rut5-base_6',]


BIO_mark_sent_v2_rut5_base_multitask_model_names = ['saved_model_BIO_mark_sent_v2_rut5-base-multitask_3',
                                                    'saved_model_BIO_mark_sent_v2_rut5-base-multitask_4',
                                                    'saved_model_BIO_mark_sent_v2_rut5-base-multitask_5',
                                                    'saved_model_BIO_mark_sent_v2_rut5-base-multitask_6',]



wb = Workbook()
# делаем единственный лист активным 
ws = wb.active

annotation = 'mark_sent_2.json'

data = load_json_data(annotation)
df = convert_to_dataframe(data)

label_list = list()
for i in range(df.shape[0]):
    label_list.extend((list(df.tags.iloc[i])))

label_list = set(label_list)


first_row = 1
count = 0
for text in texts:

  count += 1

  column_token_rubert = 1
  column_ner_rubert = 2
  column_token_rubiobert = 3
  column_ner_rubiobert = 4

  column_token_rut5 = 5
  column_ner_rut5 = 6
  column_token_rut5_multitask = 7
  column_ner_rut5_multitask = 8
  
  
  for model_name1, model_name2, model_name3, model_name4 in zip(mark_sent_v2_rubert_base_cased_model_names, 
                                                                mark_sent_v2_RuBioBERT_model_names,
                                                                mark_sent_v2_rut5_base_model_names,
                                                                mark_sent_v2_rut5_base_multitask_model_names):
    number_row = 2
    # Загрузка токенизатора и модели
    tokenizer1 = AutoTokenizer.from_pretrained(model_name1, trust_remote_code=True)
    model1 = AutoModelForTokenClassification.from_pretrained(model_name1, num_labels=len(label_list), trust_remote_code=True)

    inputs1 = tokenizer1(text, return_tensors="pt")

    with torch.no_grad():
      logits = model1(**inputs1).logits

    predictions1 = torch.argmax(logits, dim=2)
    predicted_token_class1 = [model1.config.id2label[t.item()] for t in predictions1[0]]

    tokenizer2 = AutoTokenizer.from_pretrained(model_name2, trust_remote_code=True)
    model2 = AutoModelForTokenClassification.from_pretrained(model_name2, num_labels=len(label_list), trust_remote_code=True)

    inputs2 = tokenizer2(text, return_tensors="pt")

    with torch.no_grad():
      logits = model2(**inputs2).logits

    predictions2 = torch.argmax(logits, dim=2)
    predicted_token_class2 = [model2.config.id2label[t.item()] for t in predictions2[0]]
    

    # Загрузка токенизатора и модели
    tokenizer3 = AutoTokenizer.from_pretrained(model_name3, trust_remote_code=True)
    model3 = AutoModelForTokenClassification.from_pretrained(model_name3, num_labels=len(label_list), trust_remote_code=True)

    inputs3 = tokenizer3(text, return_tensors="pt")

    with torch.no_grad():
      logits = model3(**inputs3).logits

    predictions3 = torch.argmax(logits, dim=2)
    predicted_token_class3 = [model3.config.id2label[t.item()] for t in predictions3[0]]

    tokenizer4 = AutoTokenizer.from_pretrained(model_name4, trust_remote_code=True)
    model4 = AutoModelForTokenClassification.from_pretrained(model_name4, num_labels=len(label_list), trust_remote_code=True)

    inputs4 = tokenizer4(text, return_tensors="pt")

    with torch.no_grad():
      logits = model4(**inputs4).logits

    predictions4 = torch.argmax(logits, dim=2)
    predicted_token_class4 = [model4.config.id2label[t.item()] for t in predictions4[0]]
    

    ws.cell(row=first_row, column=column_token_rubert, value=model_name1)
    ws.cell(row=first_row, column=column_token_rubiobert, value=model_name2)

    ws.cell(row=first_row, column=column_token_rut5, value=model_name3)
    ws.cell(row=first_row, column=column_token_rut5_multitask, value=model_name4)

    print('column_token_rubert =', column_token_rubert)
    print('column_ner_rubert =', column_ner_rubert)

    print('column_token_rubiobert =', column_token_rubiobert)
    print('column_ner_rubiobert =', column_ner_rubiobert)

    print('column_token_rut5 =', column_token_rut5)
    print('column_ner_rut5 =', column_ner_rut5)

    print('column_token_rut5_multitask =', column_token_rut5_multitask)
    print('column_ner_rut5_multitask =', column_ner_rut5_multitask)

    ws.merge_cells(start_row=first_row, start_column=column_token_rubert, end_row=first_row, end_column=column_ner_rubert)
    ws.merge_cells(start_row=first_row, start_column=column_token_rubiobert, end_row=first_row, end_column=column_ner_rubiobert)

    tokens1 = tokenizer1.convert_ids_to_tokens(inputs1["input_ids"][0])
    tokens2 = tokenizer2.convert_ids_to_tokens(inputs2["input_ids"][0])

    ws.merge_cells(start_row=first_row, start_column=column_token_rut5, end_row=first_row, end_column=column_ner_rut5)
    ws.merge_cells(start_row=first_row, start_column=column_token_rut5_multitask, end_row=first_row, end_column=column_ner_rut5_multitask)

    tokens3 = tokenizer3.convert_ids_to_tokens(inputs3["input_ids"][0])
    tokens4 = tokenizer4.convert_ids_to_tokens(inputs4["input_ids"][0])

    tokens1, predicted_token_class1 = merge_BERT_tokens(tokens1, predicted_token_class1)
    tokens2, predicted_token_class2 = merge_BERT_tokens(tokens2, predicted_token_class2)
    tokens3, predicted_token_class3 = merge_T5_tokens(tokens3, predicted_token_class3)
    tokens4, predicted_token_class4 = merge_T5_tokens(tokens4, predicted_token_class4)

    for source_text1, pred1, source_text2, pred2, \
        source_text3, pred3, source_text4, pred4 in zip(tokens1, predicted_token_class1, 
                                                        tokens2, predicted_token_class2,
                                                        tokens3, predicted_token_class3,
                                                        tokens4, predicted_token_class4,
                                                        ):

      ws.cell(row=number_row, column=column_token_rubert, value=source_text1)
      ws.cell(row=number_row, column=column_ner_rubert, value=pred1)
      ws.cell(row=number_row, column=column_token_rubiobert, value=source_text2)
      ws.cell(row=number_row, column=column_ner_rubiobert, value=pred2)

      ws.cell(row=number_row, column=column_token_rut5, value=source_text3)
      ws.cell(row=number_row, column=column_ner_rut5, value=pred3)
      ws.cell(row=number_row, column=column_token_rut5_multitask, value=source_text4)
      ws.cell(row=number_row, column=column_ner_rut5_multitask, value=pred4)

      number_row += 1

    column_token_rubert += 8
    column_ner_rubert += 8
    column_token_rubiobert += 8
    column_ner_rubiobert += 8

    column_token_rut5 += 8
    column_ner_rut5 += 8
    column_token_rut5_multitask += 8
    column_ner_rut5_multitask += 8

  wb.save(f'comperison_stitched_NER_model_{count}.xlsx')

print('Работа программы успешно завершина!')


