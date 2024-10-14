import json
from itertools import cycle


from transformers import AutoModelForTokenClassification, AutoTokenizer
import torch
from openpyxl import Workbook
import pandas as pd
from seqeval.metrics import accuracy_score, precision_score, recall_score, f1_score, classification_report
import matplotlib.pyplot as plt
from sklearn.metrics import roc_auc_score, roc_curve, auc
from sklearn.metrics import precision_recall_curve, average_precision_score
import numpy as np
from sklearn.preprocessing import label_binarize


# Загрузка данных из JSON
def load_json_data(filename):
    with open(filename, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data


# Преобразование данных в pandas DataFrame
def convert_to_dataframe(data):
    rows = []
    for entry in data:
        tokens = entry['tokens']
        tags = entry['tags']
        rows.append({'tokens': tokens, 'tags': tags})
    return pd.DataFrame(rows)


def joining_tokens(tokes):
    text = ' '.join(tokes)
    #text = text.replace('( ', '(').replace(' )', ')')
    #text = text.replace(' ,', ',').replace(' .', '.')
    #text = text.replace(' :', ':')
    return text

annotation = 'content/mark_sent_2.json'

# функция сшивания  токенов BERT разорваных слов 
def merge_BERT_tokens(tokens, labels):
    """
    Функция для объединения подслов в исходные слова и сохранения меток.

    tokens: Список токенов (токенизированные слова).
    labels: Список меток для каждого токена.

    Возвращает список слов с соответствующими метками.
    """
    merged_tokens = []
    merged_labels = []
    
    current_token = ""
    current_label = labels[0]
    
    for token, label in zip(tokens, labels):
        if token.startswith("##"):
            current_token += token[2:]  # Убираем префикс '##' и добавляем к текущему слову
        else:
            if current_token:  # Если есть собранное слово, добавляем его в результат
                merged_tokens.append(current_token)
                merged_labels.append(current_label)
            current_token = token  # Начинаем новое слово
            current_label = label
    
    # Добавляем последний токен
    if current_token:
        merged_tokens.append(current_token)
        merged_labels.append(current_label)
    
    return merged_tokens, merged_labels

data = load_json_data(annotation)
df = convert_to_dataframe(data)

label_list = list()
for i in range(df.shape[0]):
    label_list.extend((list(df.tags.iloc[i])))

label_list = set(label_list)

# Загрузка модели и токенизатора (предобученная на NER)
mark_sent_v2_rubert_base_cased_model_names = ['saved_model_mark_sent_2_rubert-base-cased_3',
                                              'saved_model_mark_sent_2_rubert-base-cased_4', 
                                              'saved_model_mark_sent_2_rubert-base-cased_5',
                                              'saved_model_mark_sent_2_rubert-base-cased_6',
                                              'saved_model_mark_sent_2_rubert-base-cased_7',
                                              'saved_model_mark_sent_2_rubert-base-cased_8', 
                                              'saved_model_mark_sent_2_rubert-base-cased_9',
                                              'saved_model_mark_sent_2_rubert-base-cased_10',
                                              'saved_model_mark_sent_2_rubert-base-cased_11', 
                                              'saved_model_mark_sent_2_rubert-base-cased_12',
                                              'saved_model_mark_sent_2_rubert-base-cased_13',
                                              'saved_model_mark_sent_2_rubert-base-cased_14', 
                                              'saved_model_mark_sent_2_rubert-base-cased_15',
                                              'saved_model_mark_sent_2_rubert-base-cased_16',
                                              'saved_model_mark_sent_2_rubert-base-cased_17', 
                                              'saved_model_mark_sent_2_rubert-base-cased_18',
                                              'saved_model_mark_sent_2_rubert-base-cased_19',
                                              'saved_model_mark_sent_2_rubert-base-cased_20',
                                             ]

# Пример тестовых данных

wb = Workbook()
# делаем единственный лист активным 
ws = wb.active

FIRST_ROW = 1
COLUMN_MODEL_NAME = 1
COLUMN_ACCURACY = 2
COLUMN_PRECISION = 3
COLUMN_RECALL = 4
COLUMN_F1 = 5
COLUMN_AUC = 6

ws.cell(row=FIRST_ROW, column=COLUMN_MODEL_NAME, value='Название модели')
ws.cell(row=FIRST_ROW, column=COLUMN_ACCURACY, value='Accuracy')
ws.cell(row=FIRST_ROW, column=COLUMN_PRECISION, value='Precision')
ws.cell(row=FIRST_ROW, column=COLUMN_RECALL, value='Recall')
ws.cell(row=FIRST_ROW, column=COLUMN_F1, value='F1')
ws.cell(row=FIRST_ROW, column=COLUMN_AUC, value='AUC')


test_sentences = list()
test_labels = list()
test_tokens = list()
with open('content/test_mark_text.json', 'r', encoding='utf-8') as file:
    data = json.load(file)
    for item in data:
        test_sentences.append(joining_tokens(item['tokens']))
        test_tokens.append(item['tokens'])
        test_labels.append(item['tags'])


row_number = 1

for model_name in mark_sent_v2_rubert_base_cased_model_names:

    row_number += 1

    pred_labels = list()

    tokens_list = list()

    tokenizer = AutoTokenizer.from_pretrained(model_name, trust_remote_code=True)
    model = AutoModelForTokenClassification.from_pretrained(model_name, num_labels=len(label_list), trust_remote_code=True)

    # Перевод модели в режим оценки
    model.eval()
    
    # Токенизация данных
    for test_sentence in test_sentences:
        inputs = tokenizer(test_sentence, padding=True, truncation=True, return_tensors="pt")

        # Получение предсказаний
        with torch.no_grad():
            logits = model(**inputs).logits

        predictions = torch.argmax(logits, dim=2)
        predicted_token_class = [model.config.id2label[t.item()] for t in predictions[0]]

        tokens = tokenizer.convert_ids_to_tokens(inputs["input_ids"][0])

        tokens, predicted_token_class = merge_BERT_tokens(tokens, predicted_token_class)
        tokens_list.append(tokens)
        predicted_token_class = predicted_token_class[1:-1]
        pred_labels.append(predicted_token_class)
    
    
    print('(len(test_labels) =', (len(test_labels)))
    print('(len(pred_labels) =', (len(pred_labels)))
    test_labels_cleared = list()
    pred_labels_cleared = list()
    for i in range(len(pred_labels)):
        if len(pred_labels[i]) == len(test_labels[i]):
            test_labels_cleared.append(test_labels[i])
            pred_labels_cleared.append(pred_labels[i])
    # Вычисление метрик
    accuracy = accuracy_score(test_labels_cleared, pred_labels_cleared)
    precision = precision_score(test_labels_cleared, pred_labels_cleared)
    recall = recall_score(test_labels_cleared, pred_labels_cleared)
    f1 = f1_score(test_labels_cleared, pred_labels_cleared)

    print(f"Точность (Accuracy): {accuracy}")
    print(f"Точность (Precision): {precision}")
    print(f"Полнота (Recall): {recall}")
    print(f"F1-Score: {f1}")
    ws.cell(row=row_number, column=COLUMN_MODEL_NAME, value=model_name)
    ws.cell(row=row_number, column=COLUMN_ACCURACY, value=accuracy)
    ws.cell(row=row_number, column=COLUMN_PRECISION, value=precision)
    ws.cell(row=row_number, column=COLUMN_RECALL, value=recall)
    ws.cell(row=row_number, column=COLUMN_F1, value=f1)


    # Построение графиков метрик
    metrics = ['Accuracy', 'Precision', 'Recall', 'F1']
    values = [accuracy, precision, recall, f1]

    plt.figure(figsize=(8, 5))
    plt.bar(metrics, values, color=['blue', 'orange', 'green'])
    plt.ylim(0, 1)
    plt.title('NER Metrics for ruBERT Model')
    plt.xlabel('Metrics')
    plt.ylabel('Score')

    # Сохранение графика
    plt.savefig(f'content\\ner_metrics_{model_name}.png')

    # Бинаризация меток для вычисления AUC
    all_labels = sorted(list(set(label for sublist in test_labels_cleared for label in sublist)))
    y_true = label_binarize([label for sublist in test_labels_cleared for label in sublist], classes=all_labels)
    y_pred = label_binarize([label for sublist in pred_labels_cleared for label in sublist], classes=all_labels)

    # AUC для каждого класса и усредненный AUC
    fpr = dict()
    tpr = dict()
    roc_auc = dict()

    for i, label in enumerate(all_labels):
        fpr[i], tpr[i], _ = roc_curve(y_true[:, i], y_pred[:, i])
        roc_auc[i] = auc(fpr[i], tpr[i])

    # Средний AUC
    all_fpr = np.unique(np.concatenate([fpr[i] for i in range(len(all_labels))]))
    mean_tpr = np.zeros_like(all_fpr)

    for i in range(len(all_labels)):
        mean_tpr += np.interp(all_fpr, fpr[i], tpr[i])

    mean_tpr /= len(all_labels)

    fpr["macro"] = all_fpr
    tpr["macro"] = mean_tpr
    roc_auc["macro"] = auc(fpr["macro"], tpr["macro"])

    print(f"Точность (Precision): {precision}")
    print(f"Полнота (Recall): {recall}")
    print(f"Средний AUC: {roc_auc['macro']}")

    # Построение графиков ROC-AUC
    plt.figure(figsize=(8, 6))
    colors = cycle(["blue", "orange", "green", "red", "purple"])
    for i, color in zip(range(len(all_labels)), colors):
        plt.plot(fpr[i], tpr[i], color=color, lw=2,
                label=f'ROC curve of class {all_labels[i]} (area = {roc_auc[i]:.2f})')

    plt.plot(fpr["macro"], tpr["macro"], color="navy", linestyle="--",
            label=f'macro-average ROC curve (area = {roc_auc["macro"]:.2f})', lw=2)

    plt.plot([0, 1], [0, 1], "k--", lw=2)
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel("False Positive Rate")
    plt.ylabel("True Positive Rate")
    plt.title("ROC-AUC for NER with ruBERT")
    plt.legend(loc="lower right")

    # Сохранение графика
    plt.savefig(f'content\\ner_roc_auc_{model_name}.png')

    ws.cell(row=row_number, column=COLUMN_AUC, value=roc_auc['macro'])

    # Построение Precision-Recall кривой для каждого класса
    precision_curve = dict()
    recall_curve = dict()
    average_precision = dict()

    for i, label in enumerate(all_labels):
        precision_curve[i], recall_curve[i], _ = precision_recall_curve(y_true[:, i], y_pred[:, i])
        average_precision[i] = average_precision_score(y_true[:, i], y_pred[:, i])

    # Средняя Precision-Recall кривая (macro-average)
    precision_curve["macro"], recall_curve["macro"], _ = precision_recall_curve(y_true.ravel(), y_pred.ravel())
    average_precision["macro"] = average_precision_score(y_true, y_pred, average="macro")

    print(f"Точность (Precision): {precision}")
    print(f"Полнота (Recall): {recall}")
    print(f"Средний Average Precision: {average_precision['macro']}")

    # Построение графиков Precision-Recall
    plt.figure(figsize=(8, 6))
    for i in range(len(all_labels)):
        plt.plot(recall_curve[i], precision_curve[i], lw=2,
                label=f'Precision-Recall curve of class {all_labels[i]} (AP = {average_precision[i]:.2f})')

    plt.plot(recall_curve["macro"], precision_curve["macro"], color="navy", linestyle="--",
            label=f'macro-average Precision-Recall curve (AP = {average_precision["macro"]:.2f})', lw=2)

    plt.xlabel("Recall")
    plt.ylabel("Precision")
    plt.title("Precision-Recall curve for NER with ruBERT")
    plt.legend(loc="lower left")

    # Сохранение графика
    plt.savefig(f'content\\ner_precision_recall_{model_name}.png')

wb.save(f'content\\Сравнение оценок качества работы моделей ruBERT NER.xlsx')

print('Работа программы успешно завершина!')