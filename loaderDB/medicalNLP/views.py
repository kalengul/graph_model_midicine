from django.shortcuts import render
from django.http import HttpResponse

import os

import openpyxl
from openpyxl import load_workbook

from loader_app import models
from . import text_processing
from . import vectorization_models
from . import semantic_network

# просто метод-представление для представления веб-сервиса пользователю
def medical_index(request):
    return HttpResponse('<h1>Веб-приложение для решения задач обработки текстов на естьсвенном языке для медицины.</h1>')

# метод-представление для обучения языковой модели
def train_vectorization_models(request):
    # список интересующих загаловком инструкций, содержание которых берётся из БД
    headlineList = [
        #'ИНСТРУКЦИЯ',
        'СОСТАВ',
        'ФАРМАКОТЕРАПЕВТИЧЕСКАЯ ГРУППА',
        'ОПИСАНИЕ',
        'ЛЕКАРСТВЕННАЯ ФОРМА',
        'ДЕЙСТВУЮЩЕЕ ВЕЩЕСТВО',
        'ФАРМАКОЛОГИЧЕСКИЕ СВОЙСТВА',
        'ПОКАЗАНИЯ К ПРИМЕНЕНИЮ',
        'ПРОТИВОПОКАЗАНИЯ',
        'ПРИМЕНЕНИЕ ПРИ БЕРЕМЕННОСТИ В ПЕРИОД ГРУДНОГО ВСКАРМЛИВАНИЯ, ВЛИЯНИЕ НА ФЕРТИЛЬНОСТЬ, РЕКОМЕНДАЦИИ ДЛЯ ПАЦИЕНТОВ ДЕТОРОДНОГО ВОЗРАСТО',
        'ПРИМЕНЕНИЕ ПРИ БЕРЕМЕННОСТИ В ПЕРИОД ГРУДНОГО ВСКАРМЛИВАНИЯ, ВЛИЯНИЕ НА ФЕРТИЛЬНОСТЬ, РЕКОМЕНДАЦИИ ДЛЯ ПАЦИЕНТОВ С РЕПРОДУКТИВНЫМ ПОТЕНЦИАЛОМ',
        'СПОСОБ ПРИМЕНЕНИЯ И ДОЗЫ',
        'СПОСОБ ВВЕДЕНИЯ',
        'ПОБОЧНЫЕ ЭФФЕКТЫ',
        'ПОБОЧНЫЕ ДЕЙСТВИЯ',
        'ПЕРЕДОЗИРОВКА',
        'ВЗАИМОДЕЙСТВИЕ С ДРУГИМИ ЛЕКАРСТВЕННЫМИ ПРЕПАРАТАМИ И ДРУГИЕ ФОРМЫ ВЗАИМОДЕЙСТВИЯ',
        'ВЗАИМОДЕЙСТВИЕ С ДРУГИМИ ЛЕКАРСТВЕННЫМИ  СРЕДСТВАМИ',
        'ВЗАИМОДЕЙСТВИЕ С ПИЩЕЙ',
        'КОРРЕКЦИЯ ДОЗЫ',
        'ОСОБЫЕ УКАЗАНИЯ',
        'ФОРМА ВЫПУСКА',
        'УСЛОВИЯ ХРАНЕНИЯ',
        'СРОК ГОДНОСТИ',
        'УСЛОВИЯ ОТПУСКА ИЗ АПТЕК',
        'УСЛОВИЯ ОТПУСКА',
        'РЕЗУЛЬТАТЫ КЛИНИЧЕСКИХ ИСПЫТАНИЙ',
        #'ЮРИДИЧЕСКОЕ ЛИЦО НА ИМЯ КОТОРОГО ВЫДАНО РЕГИСТРАЦИОННОЕ УДОСТОВЕРЕНИЕ',
        #'ПРОИЗВОДИТЕЛЬ',
        #'ПРЕТЕНЗИИ ПОТРЕБИТЕЛЕЙ НАПРАВЛЯТЬ ПО АДРЕСУ'
    ]

    '''
    Загрузка списка интересующих ЛС.
    В цикле для каждого ЛС находится текст инструкций из БД.
    Получений текст подвергается предобработке с помощью метод из модуля text_processing.
    Обработаный текст предается методу обучения языковой модели из мрдуля vectorization_models.
    '''
    PATH = 'D:\\The job\\loaderDB\\loaderDB\\loader_app\\drugs files\\The list of essential medicines.txt'
    file = open(PATH, 'r', encoding='utf-8')
    drugName = file.readline()
    common_word_tokens = list()
    count_drugName = 0
    print('Загрузка токенов для обучения модели')
    while drugName:
        count_drugName += 1
        print(f'прочитано {count_drugName} наименований ЛС')
        drugINs = models.InternationalName.objects.filter(internationalName = drugName.rstrip('\n'))
        for drugIN in drugINs:
            tnId = -1
            try:
                tnId = drugIN.tradename.id
            except:
                continue
            tn = models.TradeName.objects.get(id=tnId)
            instructionText = tn.instructiontext_set.all()
            if instructionText.count() != 0:
                text = ''
                for it in instructionText:
                    if any(it.headline.upper() in hl for hl in headlineList):
                        if it:
                            textBuilder = text_processing.TextBuilder(it.content)
                            digits = '0123456789'
                            text += textBuilder.set_lower().replace_yo().removing_punctuation(digits).get_result()
                            text += ' '
                #print(f'text = {text}')
                if text:
                    word_tokens = text_processing.tokenization(text)
                    word_tokens = text_processing.removing_SW(word_tokens)
                    #word_tokens = text_processing.removing_digits(word_tokens)
                    #word_tokens = text_processing.removing_latin_literals(word_tokens)
                    word_tokens = text_processing.lemmatization(word_tokens)
                    common_word_tokens.append(word_tokens)
                    
        drugName = file.readline()
    file.close()

    print('Обучение модели началось')
    window = 10                                 # размер окна обзора
    sg = 1                                      # переменная говорящая, что аглоритм обучения модели SkipGram
    vector_size = 600                           # размерность вектора слова
    resultCode = vectorization_models.train_word2vec(tokens=common_word_tokens, window=window, vector_size=vector_size, sg=sg)

    if resultCode == 1:
        print('Обучение модели завершилось успешно!')
        return HttpResponse('<h1>Обучение модели word2vec выполнено успешно!</h1>')
    else:
        print('Что-то пошло не так!')
        return HttpResponse('Что-то пошло не так!')
    
# метод-представление для строительства семантического графа (семантической сети)
def creating_semantic_network(request):
    # загрузка списка терминов
    input_excel_file = r'D:\The job\loaderDB\loaderDB\Частотный словарь медицинских терминов.xlsx'
    wb = load_workbook(input_excel_file)        # объект книги в excel-файле
    sheet = wb[wb.sheetnames[0]]                # объект листа в книги

    ROW_NUMBER = 2                              # первой строки с медицинским термином
    A_COLUMN = 1                                # номер колонки с термином

    termList = list()                           # список основных препаратов
    for i in range(ROW_NUMBER, sheet.max_row+1):
        term = sheet.cell(row=i, column=A_COLUMN).value
        termList.append(term)

    # загрузка списка лекарств
    PATH = 'D:\\The job\\loaderDB\\loaderDB\\loader_app\\drugs files\\The list of essential medicines.txt'
    essential_medicines_list = list()           # список основных препаратов
    file = open(PATH, 'r', encoding='utf-8')
    drugName = file.readline()
    while drugName:
        essential_medicines_list.append(drugName.lower().rstrip('\n'))
        drugName = file.readline()
    file.close()

    # создание семантического графа
    threshold = 0.701   # порог семантической близости, по нему определяется будут ли связаны термены, или нет 
    resultCode = semantic_network.creating_semantic_network(essential_medicines_list, termList, threshold)

    if resultCode == 1:
        print('Создание семантического графа завершилось успешно!')
        return HttpResponse('<h1>Создание семантического графа завершилось успешно!</h1>')
    else:
        print('Что-то пошло не так!')
        return HttpResponse('Что-то пошло не так!')

        
# метод-представление для строительства сети Байеса
def creating_Bayesian_network(request):
    return HttpResponse('<h1>Создание сети Байеса!</h1>')
    