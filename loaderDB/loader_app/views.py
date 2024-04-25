from django.shortcuts import render
from django.http import HttpResponse
from django.core.paginator import Paginator

import os

from openpyxl import Workbook, load_workbook

from . import models

from .back_loader import loader
from medicalNLP import text_processing


'''
Используемые сокращения:
МНН - международные наименования
ТН - торговое наименования
'''

# переменные для сохранения коллекции с МНН и инструкциями
essentialMedicinesText = None           # МНН + тексты инструкций
essentialMedicinesLink = None           # МНН + сслыки на инструкции

def index(request):
    return HttpResponse("<h2>Это веб-приложение для получения инфомации о лекарственных средствах</h2>")

# метод-представление для загрузки данных о ЛС в БД 
def back_loader(request):
    row = loader()
    data = {'row': row}
    return render(request, 'back_loader.html', context=data)

'''
Метод-представление для загрузки данных о ЛС из БД в клиента в виде таблицы.
Формат таблицы: <МНН, Держатель, Форма выпуска>.
Таблица в соответсвующем шаблоне.
'''
def front_loader(request):
    drugNumber = models.InternationalName.objects.all().count()
    row = f'На сайте есть информация о {drugNumber} ЛС'
    head_cols = ['Название лекарства', 'Производитель', 'Формы выпуска']
    if request.method == 'POST':
        drugName = request.POST.get('drug_name', 'неопределённое имя')
        INs = models.InternationalName.objects.filter(internationalName__contains=drugName)
        if not INs:
            return HttpResponse(f"<h2 align=\"center\">Информация об указанном ЛС не найдена</h2>")
        drugs = list()
        for IN in INs:
            interName = IN.internationalName
            drug_id = IN.id
            DI = IN.druginformation
            holder = ''
            holderName = 'информации о владельце нет'
            holderCountry = 'информации о стране нет'
            if DI.holder.holderName:
                holderName = DI.holder.holderName
            if DI.holder.holderCountry:
                holderCountry = DI.holder.holderCountry
            holder = holderName + ', ' + holderCountry
            forms = ''
            for form in DI.releaseform_set.all():
                forms += form.dosageForm + '\n'
                for package in form.package_set.all():
                    forms += package.packageDescription + '\n'
            if not forms:
                forms = 'информации нет'

            drug = {
                    'name': interName,
                    'holder': holder,
                    'forms': forms,
                    'drug_id': drug_id
                    }
            drugs.append(drug)

        data = {
                'row': row,
                'head_cols': head_cols, 
                'drugs': drugs
                }
        
    else:
         data = {'row': row}
    return render(request, 'front_loader.html', context=data)

'''
Метод-предстваление для вывода таблицы с текста инструкций.
Формат таблицы: <Заголовок, Содержание под заголовком>.
Таблица в соответсвующем шаблоне.
'''
def show_instruction(request, drug_id):
    drugIN = models.InternationalName.objects.get(id=drug_id)
    drugTN_id = drugIN.tradename.id
    drugTN = models.TradeName.objects.get(id=drugTN_id)
    hc_dict_list = list()
    for it in drugTN.instructiontext_set.all():
        hc_dict = { 
                    'headline': it.headline,
                    'content': it.content
                    }
        hc_dict_list.append(hc_dict)
    head_cols = ['Заголовок', 'Содержание']
    if hc_dict_list:
        data = {
                'drug_name': drugIN.internationalName,
                'head_cols': head_cols,
                'hc_dict_list': hc_dict_list
                }
        return render(request, 'show_instruction.html', context=data)
    else:
        return HttpResponse(f'<h2 align=\"center\">Текст инструкции отсутсвует!</h2>')

'''
Метод-предстваление для вывода таблицы с МНН и текстами инструкций.
Формат таблицы: <МНН, Заголовок, Содержание под заголовком>.
Таблица в соответсвующем шаблоне.
'''
def show_list_of_essential_medicines_text (request):
    global essentialMedicinesText
    head_cols = ['Международное наименование', 'Инструкция']

    # функция для загрузки МНН и инструкция для лекарств из списка основных лекарств
    def get_essential_medicines_text():
        PATH = 'D:\\The job\\loaderDB\\loaderDB\\loader_app\\drugs files\\The list of essential medicines.txt'
        f = open(PATH, 'r', encoding='utf-8')
        ni_dict_list = list()
        drugName = f.readline()
        while drugName:
            drugINs = models.InternationalName.objects.filter(internationalName__contains=drugName.rstrip('\n'))
            for drugIN in drugINs:
                name = drugIN.internationalName
                instructions = list()
                drugTN_id = None
                try:
                    drugTN_id = drugIN.tradename.id
                except:
                    continue
                drugTN = models.TradeName.objects.get(id=drugTN_id)
                if not drugTN.instructiontext_set.all():
                    continue
                for it in drugTN.instructiontext_set.all():
                    instruction = { 
                                'headline': it.headline,
                                'content': it.content
                                }
                    instructions.append(instruction)
                ni_dict = {
                            'name': name,
                            'instructions': instructions
                            }
                ni_dict_list.append(ni_dict)
            drugName = f.readline()
        f.close()
        return ni_dict_list

    # если основные лекарства не загружены в коллекцию, загружаем
    if not essentialMedicinesText:
        essentialMedicinesText = get_essential_medicines_text()

    # пагинация для пролистования результатов запроса
    paginator = Paginator(essentialMedicinesText, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    data = {
            'head_cols': head_cols,
            'page_obj': page_obj
            }
    
    return render(request, 'show_list_of_essential_medicines_text.html', context=data)

'''
Метод-предстваление для вывода таблицы с МНН и ссылками на инструкции.
Формат таблицы: <МНН, Список ссылкол на инструкции>.
Таблица в соответсвующем шаблоне.
'''
def show_list_of_essential_medicines_link(request):
    global essentialMedicinesLink
    head_cols = ['Международное наименование', 'Инструкции']

    # функция для загрузки МНН и инструкция для лекарств из списка основных лекарств
    def get_essential_medicines_link():
        PATH = 'D:\\The job\\loaderDB\\loaderDB\\loader_app\\drugs files\\The list of essential medicines.txt'
        f = open(PATH, 'r', encoding='utf-8')
        ni_dict_list = list()
        drugName = f.readline()
        while drugName:
            drugINs = models.InternationalName.objects.filter(internationalName__contains=drugName.rstrip('\n'))
            nameList = list()
            for drugIN in drugINs:
                nameList.append(drugIN.internationalName)
            nameListSet = set(nameList)
            for uniqueDrugIN in nameListSet:
                name = uniqueDrugIN
                instructions = list()
                for drugIN in drugINs:
                    if uniqueDrugIN == drugIN.internationalName:
                        drugTN_name = None
                        try:
                            drugTN_name = drugIN.tradename.tradeName
                        except:
                            continue
                        drugInfo_id = None
                        year = 'год выпуска не указан'
                        try:
                            drugInfo_id = drugIN.druginformation.id
                        except:
                            continue
                        try: 
                            year = drugIN.druginformation.stateRegistrationDate
                        except:
                            print('какие-то проблемы с годом!')
                        if not year:
                            year = 'год выпуска не указан'
                        drugInfo = models.DrugInformation.objects.get(id=drugInfo_id)
                        if not drugInfo.druginstruction_set.all():
                            continue
                        URLs = list()
                        for URL in drugInfo.druginstruction_set.all():
                            if 'http' in URL.URL:
                                URLs.append(URL.URL)
                        if URLs:
                            instruction = { 
                                    'tn': drugTN_name,
                                    'year': year,
                                    'URLs': URLs
                                    }
                            instructions.append(instruction)
                ni_dict = {
                            'name': name,
                            'instructions': instructions
                            }
                ni_dict_list.append(ni_dict)
            drugName = f.readline()
        f.close()
        return ni_dict_list
    

    # если основные лекарства не загружены в коллекцию, загружаем
    if not essentialMedicinesLink:
        essentialMedicinesLink = get_essential_medicines_link()

    # пагинация для пролистования результатов запроса
    paginator = Paginator(essentialMedicinesLink, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    data = {
            'head_cols': head_cols,
            'page_obj': page_obj
            }
    
    return render(request, 'show_list_of_essential_medicines_link.html', context=data)

'''
Метод-предстваление для вывода таблицы с МНН и ссылками на инструкции в excel-файл.
Формат таблицы: <МНН, ТН, Год, Список ссылкол на инструкции>.
В шаблоне просто выводится сообщении о завершении работы метода.
'''
def load_to_excel(request):

    head_cols = ['Международное наименование', 'Инструкции']

    def get_essential_medicines_link():
        PATH = 'D:\\The job\\loaderDB\\loaderDB\\loader_app\\drugs files\\The list of essential medicines.txt'
        f = open(PATH, 'r', encoding='utf-8')
        ni_dict_list = list()
        drugName = f.readline()
        while drugName:
            drugINs = models.InternationalName.objects.filter(internationalName__contains=drugName.rstrip('\n'))
            nameList = list()
            for drugIN in drugINs:
                nameList.append(drugIN.internationalName)
            nameListSet = set(nameList)
            for uniqueDrugIN in nameListSet:
                name = uniqueDrugIN
                instructions = list()
                for drugIN in drugINs:
                    if uniqueDrugIN == drugIN.internationalName:
                        drugTN_name = None
                        try:
                            drugTN_name = drugIN.tradename.tradeName
                        except:
                            continue
                        drugInfo_id = None
                        year = 'год выпуска не указан'
                        try:
                            drugInfo_id = drugIN.druginformation.id
                        except:
                            continue
                        try: 
                            year = drugIN.druginformation.stateRegistrationDate
                        except:
                            print('какие-то проблемы с годом!')
                        if not year:
                            year = 'год выпуска не указан'
                        drugInfo = models.DrugInformation.objects.get(id=drugInfo_id)
                        if not drugInfo.druginstruction_set.all():
                            continue
                        URLs = list()
                        for URL in drugInfo.druginstruction_set.all():
                            if 'http' in URL.URL:
                                URLs.append(URL.URL)
                        if URLs:
                            instruction = { 
                                    'tn': drugTN_name,
                                    'year': year,
                                    'URLs': URLs
                                    }
                            instructions.append(instruction)
                ni_dict = {
                            'name': name,
                            'instructions': instructions
                            }
                ni_dict_list.append(ni_dict)
            drugName = f.readline()
        f.close()
        return ni_dict_list

    
    print('Загрузка данных из БД')
    ni_dict_list = get_essential_medicines_link()

    '''
    ni_dict_list = [
                        {
                            'name': 'МНН1',
                            'instructions': 
                            [
                                {
                                    'tn': 'ТН1',
                                    'year': 'год1',
                                    'URLs': ['URL1', 'URL2']
                                },
                                {
                                    'tn': 'ТН2',
                                    'year': 'год2',
                                    'URLs': ['URL1', 'URL2']
                                }
                            ]
                        },
                        {
                            'name': 'МНН2',
                            'instructions': 
                            [
                                {
                                    'tn': 'ТН1',
                                    'year': 'год1',
                                    'URLs': ['URL1', 'URL2']
                                },
                                {
                                    'tn': 'ТН2',
                                    'year': 'год2',
                                    'URLs': ['URL1', 'URL2']
                                },
                                {
                                    'tn': 'ТН3',
                                    'year': 'год2',
                                    'URLs': ['URL1', 'URL2', 'URL3']
                                }
                            ]
                        }
                    ]
    '''
    print('Загрузка данных в excel-файл')
    wb = Workbook()

    sheet = wb.active
    rowHeadline_num = 1
    colA_num = 1
    colB_num = 2
    colC_num = 3
    colD_num = 4
    sheet.merge_cells(start_row=rowHeadline_num, start_column=colB_num, end_row=rowHeadline_num, end_column=colD_num)
    sheet.cell(row=1, column=1, value=head_cols[0])
    sheet.cell(row=1, column=2, value=head_cols[1])
    row_num_IN = 2
    row_num_TN = 2
    row_num_URL = 2

    for item in ni_dict_list:
        URL_count = 0
        for i in item['instructions']:
            increment = len(i['URLs'])
            URL_count += increment
        if URL_count == 0:
            continue
        for i in item['instructions']:
            increment = len(i['URLs'])
            if increment == 0:
                continue
            for url in i['URLs']:
                sheet.cell(row=row_num_URL, column=colD_num, value=url)
                row_num_URL += 1
            print(f'increment = {increment}')
            end_row = row_num_TN + increment - 1
            print(f'row_num_TN = {row_num_TN}')
            print(f'end_row = {end_row}')
            sheet.merge_cells(start_row=row_num_TN, start_column=colB_num, end_row=end_row, end_column=colB_num)
            sheet.cell(row=row_num_TN, column=colB_num, value=i['tn'])
            sheet.merge_cells(start_row=row_num_TN, start_column=colC_num, end_row=end_row, end_column=colC_num)
            sheet.cell(row=row_num_TN, column=colC_num, value=i['year'])
            row_num_TN += increment

        print('URL_count = ', URL_count)
        end_IN_row = row_num_IN + URL_count - 1
        print('row_num_IN =', row_num_IN)
        print('end_IN_row =', end_IN_row)
        print(row_num_IN < end_IN_row)
        sheet.merge_cells(start_row=row_num_IN, start_column=colA_num, end_row=end_IN_row, end_column=colA_num)
        sheet.cell(row=row_num_IN, column=colA_num, value=item['name'])
        row_num_IN += URL_count
        wb.save('output_excel.xlsx')
       
    print('Работа функции по выгрузки данных о лекарственных средствах завершена успешно!')
        
    return HttpResponse('<h2>Выгрузка данных о лекарственных средствах выполнена успешно!</h2>')


'''
Метод-предстваление для вывода частотного слова слов (терминов) из текстов инструкций.
Частоты слов подсчитываются для каждого ЛС отдельно.
Формат таблицы: <МНН, термин, частота термина>.
В шаблоне просто выводится сообщении о завершении работы метода.
'''
def frequency_dictionary_for_each_drug_to_excel(request):
    headlineList = [
        #'ИНСТРУКЦИЯ',
        'СОСТАВ',
        'ФАРМАКОТЕРАПЕВТИЧЕСКАЯ ГРУППА',
        'ОПИСАНИЕ',
        'ЛЕКАРСТВЕННАЯ ФОРМА',
        'ДЕЙСТВУЮЩЕЕ ВЕЩЕСТВО',
        'ФАРМАКОЛОГИЧЕСКИЕ СВОЙСТВА',
        'ПОКАЗАНИЯ К ПРИМЕНЕНИЮ',
        'ПРОТИВОПОКАЗАНИЯ',
        'ПРИМЕНЕНИЕ ПРИ БЕРЕМЕННОСТИ В ПЕРИОД ГРУДНОГО ВСКАРМЛИВАНИЯ, ВЛИЯНИЕ НА ФЕРТИЛЬНОСТЬ, РЕКОМЕНДАЦИИ ДЛЯ ПАЦИЕНТОВ ДЕТОРОДНОГО ВОЗРАСТО',
        'ПРИМЕНЕНИЕ ПРИ БЕРЕМЕННОСТИ В ПЕРИОД ГРУДНОГО ВСКАРМЛИВАНИЯ, ВЛИЯНИЕ НА ФЕРТИЛЬНОСТЬ, РЕКОМЕНДАЦИИ ДЛЯ ПАЦИЕНТОВ С РЕПРОДУКТИВНЫМ ПОТЕНЦИАЛОМ',
        'СПОСОБ ПРИМЕНЕНИЯ И ДОЗЫ',
        'СПОСОБ ВВЕДЕНИЯ',
        'ПОБОЧНЫЕ ЭФФЕКТЫ',
        'ПОБОЧНЫЕ ДЕЙСТВИЯ',
        'ПЕРЕДОЗИРОВКА',
        'ВЗАИМОДЕЙСТВИЕ С ДРУГИМИ ЛЕКАРСТВЕННЫМИ ПРЕПАРАТАМИ И ДРУГИЕ ФОРМЫ ВЗАИМОДЕЙСТВИЯ',
        'ВЗАИМОДЕЙСТВИЕ С ДРУГИМИ ЛЕКАРСТВЕННЫМИ  СРЕДСТВАМИ',
        'ВЗАИМОДЕЙСТВИЕ С ПИЩЕЙ',
        'КОРРЕКЦИЯ ДОЗЫ',
        'ОСОБЫЕ УКАЗАНИЯ',
        'ФОРМА ВЫПУСКА',
        'УСЛОВИЯ ХРАНЕНИЯ',
        'СРОК ГОДНОСТИ',
        'УСЛОВИЯ ОТПУСКА ИЗ АПТЕК',
        'УСЛОВИЯ ОТПУСКА',
        'РЕЗУЛЬТАТЫ КЛИНИЧЕСКИХ ИСПЫТАНИЙ',
        #'ЮРИДИЧЕСКОЕ ЛИЦО НА ИМЯ КОТОРОГО ВЫДАНО РЕГИСТРАЦИОННОЕ УДОСТОВЕРЕНИЕ',
        #'ПРОИЗВОДИТЕЛЬ',
        #'ПРЕТЕНЗИИ ПОТРЕБИТЕЛЕЙ НАПРАВЛЯТЬ ПО АДРЕСУ'
    ]


    wb = Workbook()
    sheet_number = 1
    LAST_SHEET = -1
    sheet_name = 'Лист' + str(sheet_number)
    sheet = wb.create_sheet(sheet_name, LAST_SHEET)
    HEADLINE_NAME = 'Международное наименование'
    HEADLINE_WORD = 'Слова'
    HEADLINE_FREQUENCY = 'Частота'
    HEAD_LINE_NUMBER  = 1
    row_number = 2
    A_column_row = 2
    increment = 0
    A_COLUMN = 1
    B_COLUMN = 2
    C_COLUMN = 3
    ROW_LIMIT = 1048576

    sheet.cell(row=HEAD_LINE_NUMBER, column=A_COLUMN, value=HEADLINE_NAME)
    sheet.cell(row=HEAD_LINE_NUMBER, column=B_COLUMN, value=HEADLINE_WORD)
    sheet.cell(row=HEAD_LINE_NUMBER, column=C_COLUMN, value=HEADLINE_FREQUENCY)
    wb.save('Частотный словарь.xlsx')
    
    PATH = 'D:\\The job\\loaderDB\\loaderDB\\loader_app\\drugs files\\The list of essential medicines.txt'
    file = open(PATH, 'r', encoding='utf-8')
    drugName = file.readline()

    while drugName:
        drugINs = models.InternationalName.objects.filter(internationalName = drugName.rstrip('\n'))
        for drugIN in drugINs:
            tnId = -1
            try:
                tnId = drugIN.tradename.id
            except:
                continue
            tn = models.TradeName.objects.get(id=tnId)
            instructionText = tn.instructiontext_set.all()
            if instructionText.count() != 0:
                text = ''
                for it in instructionText:
                    if any(it.headline.upper() in hl for hl in headlineList):
                        if it:
                            textBuilder = text_processing.TextBuilder(it.content)
                            text += textBuilder.set_lower().replace_yo().removing_punctuation().get_result()
                            text += ' '
                if text:
                    word_tokens = text_processing.tokenization(text)
                    word_tokens = text_processing.removing_SW(word_tokens)
                    word_tokens = text_processing.lemmatization(word_tokens)
                    freqDict = text_processing.frequency_dictionary(word_tokens)
                    print(f'len(freqDict) + A_column_row = {len(freqDict) + A_column_row}')
                    if (increment + A_column_row) >= ROW_LIMIT:
                        print('Новый лист!')
                        sheet_number += 1
                        sheet_name = 'Лист' + str(sheet_number)
                        sheet = wb.create_sheet(sheet_name, LAST_SHEET)
                        row_number = 2
                        A_column_row = 2
                        increment = 0
                        sheet.cell(row=HEAD_LINE_NUMBER, column=A_COLUMN, value=HEADLINE_NAME)
                        sheet.cell(row=HEAD_LINE_NUMBER, column=B_COLUMN, value=HEADLINE_WORD)
                        sheet.cell(row=HEAD_LINE_NUMBER, column=C_COLUMN, value=HEADLINE_FREQUENCY)
                    
                    A_column_row += increment
                    sheet.cell(row = A_column_row, column = A_COLUMN, value = drugIN.internationalName)
                    for w, f in freqDict.items():
                        sheet.cell(row = row_number, column = B_COLUMN, value = w)
                        sheet.cell(row = row_number, column = C_COLUMN, value = f)
                        row_number += 1
                        increment += 1
                    
        drugName = file.readline()
    wb.remove(wb['Sheet'])
    wb.save('Частотный словарь.xlsx')
    file.close()
    return HttpResponse('<h1>Частотный словарь создан успешно!</h1>')


'''
Метод-предстваление для вывода частотного слова слов (терминов) из текстов инструкций.
Частоты слов подсчитываются по всем ЛС.
Формат таблицы: <МНН, термин, частота термина>.
В шаблоне просто выводится сообщении о завершении работы метода.
'''
def frequency_dictionary_for_all_drugs_to_excel(request):
    headlineList = [
        #'ИНСТРУКЦИЯ',
        'СОСТАВ',
        'ФАРМАКОТЕРАПЕВТИЧЕСКАЯ ГРУППА',
        'ОПИСАНИЕ',
        'ЛЕКАРСТВЕННАЯ ФОРМА',
        'ДЕЙСТВУЮЩЕЕ ВЕЩЕСТВО',
        'ФАРМАКОЛОГИЧЕСКИЕ СВОЙСТВА',
        'ПОКАЗАНИЯ К ПРИМЕНЕНИЮ',
        'ПРОТИВОПОКАЗАНИЯ',
        'ПРИМЕНЕНИЕ ПРИ БЕРЕМЕННОСТИ В ПЕРИОД ГРУДНОГО ВСКАРМЛИВАНИЯ, ВЛИЯНИЕ НА ФЕРТИЛЬНОСТЬ, РЕКОМЕНДАЦИИ ДЛЯ ПАЦИЕНТОВ ДЕТОРОДНОГО ВОЗРАСТО',
        'ПРИМЕНЕНИЕ ПРИ БЕРЕМЕННОСТИ В ПЕРИОД ГРУДНОГО ВСКАРМЛИВАНИЯ, ВЛИЯНИЕ НА ФЕРТИЛЬНОСТЬ, РЕКОМЕНДАЦИИ ДЛЯ ПАЦИЕНТОВ С РЕПРОДУКТИВНЫМ ПОТЕНЦИАЛОМ',
        'СПОСОБ ПРИМЕНЕНИЯ И ДОЗЫ',
        'СПОСОБ ВВЕДЕНИЯ',
        'ПОБОЧНЫЕ ЭФФЕКТЫ',
        'ПОБОЧНЫЕ ДЕЙСТВИЯ',
        'ПЕРЕДОЗИРОВКА',
        'ВЗАИМОДЕЙСТВИЕ С ДРУГИМИ ЛЕКАРСТВЕННЫМИ ПРЕПАРАТАМИ И ДРУГИЕ ФОРМЫ ВЗАИМОДЕЙСТВИЯ',
        'ВЗАИМОДЕЙСТВИЕ С ДРУГИМИ ЛЕКАРСТВЕННЫМИ  СРЕДСТВАМИ',
        'ВЗАИМОДЕЙСТВИЕ С ПИЩЕЙ',
        'КОРРЕКЦИЯ ДОЗЫ',
        'ОСОБЫЕ УКАЗАНИЯ',
        'ФОРМА ВЫПУСКА',
        'УСЛОВИЯ ХРАНЕНИЯ',
        'СРОК ГОДНОСТИ',
        'УСЛОВИЯ ОТПУСКА ИЗ АПТЕК',
        'УСЛОВИЯ ОТПУСКА',
        'РЕЗУЛЬТАТЫ КЛИНИЧЕСКИХ ИСПЫТАНИЙ',
        #'ЮРИДИЧЕСКОЕ ЛИЦО НА ИМЯ КОТОРОГО ВЫДАНО РЕГИСТРАЦИОННОЕ УДОСТОВЕРЕНИЕ',
        #'ПРОИЗВОДИТЕЛЬ',
        #'ПРЕТЕНЗИИ ПОТРЕБИТЕЛЕЙ НАПРАВЛЯТЬ ПО АДРЕСУ'
    ]

    wb = Workbook()
    sheet_number = 1
    LAST_SHEET = -1
    sheet_name = 'Лист' + str(sheet_number)
    sheet = wb.create_sheet(sheet_name, LAST_SHEET)
    HEADLINE_NAME = 'Международное наименование'
    HEADLINE_WORD = 'Слова'
    HEADLINE_FREQUENCY = 'Частота'
    HEAD_LINE_NUMBER  = 1
    row_number = 2
    increment = 0
    A_COLUMN = 1
    B_COLUMN = 2
    ROW_LIMIT = 1048576

    sheet.cell(row=HEAD_LINE_NUMBER, column=A_COLUMN, value=HEADLINE_WORD)
    sheet.cell(row=HEAD_LINE_NUMBER, column=B_COLUMN, value=HEADLINE_FREQUENCY)
    wb.save('Частотный словарь.xlsx')
    
    print('Создание частотного словаря')
    PATH = 'D:\\The job\\loaderDB\\loaderDB\\loader_app\\drugs files\\The list of essential medicines.txt'
    file = open(PATH, 'r', encoding='utf-8')
    drugName = file.readline()
    common_word_tokens = list()
    count_drugName = 0
    while drugName:
        count_drugName += 1
        print(f'прочитано {count_drugName} наименований ЛС')
        drugINs = models.InternationalName.objects.filter(internationalName = drugName.rstrip('\n'))
        for drugIN in drugINs:
            tnId = -1
            try:
                tnId = drugIN.tradename.id
            except:
                continue
            tn = models.TradeName.objects.get(id=tnId)
            instructionText = tn.instructiontext_set.all()
            if instructionText.count() != 0:
                text = ''
                for it in instructionText:
                    if any(it.headline.upper() in hl for hl in headlineList):
                        if it:
                            textBuilder = text_processing.TextBuilder(it.content)
                            digits = '0123456789'
                            text += textBuilder.set_lower().replace_yo().removing_punctuation(digits).get_result()
                            text += ' '
                if text:
                    word_tokens = text_processing.tokenization(text)
                    word_tokens = text_processing.removing_SW(word_tokens)
                    word_tokens = text_processing.lemmatization(word_tokens)
                    common_word_tokens += word_tokens
                    
        drugName = file.readline()
    file.close()

    print('Выгрузак частотного словаря в excel-файл')
    freqDict = text_processing.frequency_dictionary(common_word_tokens)
    word_count = 0
    for w, f in freqDict.items():
        word_count += 1
        if word_count % 10000 == 0:
            print(f'записано {word_count} слов')
        if row_number >= ROW_LIMIT:
            print('Новый лист!')
            sheet_number += 1
            sheet_name = 'Лист' + str(sheet_number)
            sheet = wb.create_sheet(sheet_name, LAST_SHEET)
            row_number = 2
            sheet.cell(row=HEAD_LINE_NUMBER, column=A_COLUMN, value=HEADLINE_WORD)
            sheet.cell(row=HEAD_LINE_NUMBER, column=B_COLUMN, value=HEADLINE_FREQUENCY)
        
        sheet.cell(row = row_number, column = A_COLUMN, value = w)
        sheet.cell(row = row_number, column = B_COLUMN, value = f)
        row_number += 1
    wb.remove(wb['Sheet'])
    wb.save('Частотный словарь.xlsx')

    return HttpResponse('<h1>Частотный словарь создан успешно!</h1>')

'''
Метод-предстваление для вывода отфильтрованного частотного слова слов (терминов) из текстов инструкций.
Отфильтрованного частоный словарь содержит только медицинские термины и их частоты.
Формат таблицы: <МНН, термин, частота термина>.
В шаблоне просто выводится сообщении о завершении работы метода.
'''
def filtering_medical_terms(request):
    INPUT_EXCEL_FILENAME = 'Частотный словарь.xlsx'
    wb = load_workbook(INPUT_EXCEL_FILENAME)
    sheet = wb['Лист1']
    ROW_NUMBER = 2
    A_COLUMN = 1
    B_COLUMN = 2
    word_freq_dict = dict()
    for i in range(ROW_NUMBER, sheet.max_row+1):
        word_freq_dict[sheet.cell(row = i, column = A_COLUMN).value] = sheet.cell(row = i, column = B_COLUMN).value
    word_list = list()
    DIR_PATH = 'D:\\The job\\loaderDB\\loaderDB\\russian words'
    for file_name in os.listdir(DIR_PATH):
        file_path = os.path.join(DIR_PATH, file_name)
        file = open(file_path, 'r', encoding='utf-8')
        word = file.readline().rstrip('\n')
        if word:
            word_list.append(word)
        while word:
            word = file.readline().rstrip('\n')
            word_list.append(word)
        file.close()
    
    word_set = set(word_list)
    print(f'len(word_set) = {len(word_set)}')
    words_for_excel = set(word_freq_dict.keys())
    print(f'len(words_for_excel) = {len(words_for_excel)}')
    medical_terms = words_for_excel - word_set
    print(f'len(medical_terms) = {len(medical_terms)}')
    medical_terms_freq = dict()
    for medical_term in medical_terms:
        medical_terms_freq[medical_term] = word_freq_dict[medical_term]
    print(f'len(medical_terms_freq) = {len(medical_terms_freq)}')

    wb = Workbook()
    sheet = wb.active
    OUTPUT_EXCEL_FILENAME = 'Частотный словарь медицинских терминов.xlsx'
    HEADLINE_WORD = 'Слова'
    HEADLINE_FREQUENCY = 'Частота'
    HEAD_LINE_NUMBER  = 1
    row_number = 2

    sheet.cell(row=HEAD_LINE_NUMBER, column=A_COLUMN, value=HEADLINE_WORD)
    sheet.cell(row=HEAD_LINE_NUMBER, column=B_COLUMN, value=HEADLINE_FREQUENCY)
    wb.save(OUTPUT_EXCEL_FILENAME)

    for k, v in medical_terms_freq.items():
        sheet.cell(row=row_number, column=A_COLUMN, value=k)
        sheet.cell(row=row_number, column=B_COLUMN, value=v)
        row_number += 1
    wb.save(OUTPUT_EXCEL_FILENAME)

    return HttpResponse('<h1>Фильтрация медицинских терминов выполнена успешно!</h1>')