from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

import os
from parse_html import drug

from googletrans import Translator
import time

translator = Translator()

def translate_text(text):
    while True:
        try:
            translation = translator.translate(text, dest='en')
            return translation.text
        except Exception as e:
            print(f"Ошибка перевода: {e}")
            time.sleep(0.5)  # Ждем перед повторной попыткой

def set_bold_text(sheet, row, column, text):
    cell = f"{get_column_letter(column)}{row}"
    sheet[cell] = text
    sheet[cell].font = sheet[cell].font.copy(bold=True)

# Получаем список файлов в папке
files = os.listdir('./drugs_html')


drug_list = []
for file_name in files:
    drug_list.append(drug(file_name))

# Создаем словарь для хранения уникальных подпобочных элементов для каждого побочного эффекта
unique_sub_side_effects = {}

# Для каждого побочного эффекта собираем уникальные подпобочные элементы
for drug_element in drug_list:
    for effect in drug_element.drug_side_effect_list:
        if effect.name not in unique_sub_side_effects:
            unique_sub_side_effects[effect.name] = set()
        for sub_effect in effect.sub_side_effect:
            unique_sub_side_effects[effect.name].add(sub_effect)


for side_effect in unique_sub_side_effects:
    print(f"Побочный эффект: {side_effect}")

    for sub_side_effect in unique_sub_side_effects[side_effect]:
        print(f"Подбочный эффект: {sub_side_effect}")

# Возрастные группы
age_list = [
    "0 - 27 days",
    "28 days to 23 months",
    "2 - 11 years",
    "12 - 17 years",
    "18 - 44 years",
    "45 - 64 years",
    "65 - 74 years",
    "≥ 75 years",
    "Unknown"
]

# Пол
sex_list = ["Female", "Male", "Unknown"]

# # Начальные строки и столбцы
# start_row_age = 5
# start_row_sex = 12

# Создание новой книги Excel
wb = Workbook()
sheet = wb.active

sheet.cell(row=1, column=1, value="Имя файла")              # Записываем названия классов в первую строку
sheet.cell(row=2, column=1, value="Название ru")            # Записываем названия классов в первую строку
sheet.cell(row=3, column=1, value="Название en")            # Записываем названия классов в первую строку
sheet.cell(row=4, column=1, value="Количество обращений")   # Записываем названия классов в первую строку

# Запись данных о возрасте
for row_offset, age_name in enumerate(age_list):
    sheet.cell(row=5 + row_offset, column=1, value=age_name)

# Запись данных о поле
for row_offset, sex_name in enumerate(sex_list):
    sheet.cell(row=14 + row_offset, column=1, value=sex_name)

offset = 17

# Выводим уникальные подпобочные эффекты для каждого побочного эффекта, отсортированные по алфавиту
for side_effect in sorted(unique_sub_side_effects):

    set_bold_text(sheet, offset, 1, translate_text(side_effect))
    offset += 1
    # print("offset:", offset)

    for sub_effect in sorted(unique_sub_side_effects[side_effect]):
        # sheet.cell(row=offset, column=1, value= translate_text(sub_effect))
        sheet.cell(row=offset, column=1, value= sub_effect)
        offset += 1
        # print("offset:", offset)


# counter = 0

# Запись данных из массива классов в столбцы
# max_length = max(drug.get_len_list() for drug in drug_list)  # Находим максимальную длину массива данных
for i, drg in enumerate(drug_list):
    print(drg.name_ru)
    sheet.cell(row=1, column=i+2, value=drg.file_name)  # Записываем названия классов в первую строку
    sheet.cell(row=2, column=i+2, value=drg.name_ru)    # Записываем названия классов в первую строку
    sheet.cell(row=3, column=i+2, value=drg.name_en)    # Записываем названия классов в первую строку
    sheet.cell(row=4, column=i+2, value=drg.number)     # Записываем названия классов в первую строку

    # Запись данных о возрасте
    for row_offset, age_name in enumerate(age_list):
        count_age = drg.drug_age_list.get(age_name, 0)
        sheet.cell(row=5 + row_offset, column=i+2, value=count_age)

    # Запись данных о поле
    for row_offset, sex_name in enumerate(sex_list):
        count_sex = drg.drug_sex_list.get(sex_name, 0)
        sheet.cell(row=14 + row_offset, column=i+2, value=count_sex)

    offset = 17
    counter = 0
    # Цикл будет выполняться, пока значение ячейки не станет пустым
    while sheet.cell(row=offset, column=1).value is not None:
        effect_value = sheet.cell(row=offset, column=1).value
        # print("effect_value", offset, effect_value)
        value = 0
        flag = 0

        # drug_side_effect = drg.drug_side_effect_list[counter]
        # if drug_side_effect.name == effect_value:
        #     value = drug_side_effect.cases
        #     flag = 1
        #     counter += 1
        #     break

        # value = drug_side_effect.sub_side_effect.get(effect_value, 0)

        # Проходим по списку побочных эффектов
        for drug_side_effect in drg.drug_side_effect_list:
            if drug_side_effect.name == effect_value:
                value = drug_side_effect.cases
                flag = 1
            else:
                value = drug_side_effect.sub_side_effect.get(effect_value, 0)

            if value != 0:
                break

                


        # # Проходим по списку побочных эффектов
        # for drug_side_effect in drg.drug_side_effect_list:
        #     if drug_side_effect.name == effect_value:
        #         value = drug_side_effect.cases
        #         flag = 1
        #         counter += 1
        #         break
        #     for sub_side_effect in drug_side_effect.sub_side_effect:
        #         if sub_side_effect['name'] == effect_value:
        #             value = sub_side_effect['cases']
        #             break
    
        if flag == 1:
            set_bold_text(sheet, offset, i+2, value)
        else:
            sheet.cell(row=offset, column=i+2, value=value)  # Записываем данные в соответствующий столбец

        offset += 1


# Сохраняем книгу
wb.save('test_3.xlsx')