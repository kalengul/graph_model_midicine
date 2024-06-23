import difflib

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def similarity_percentage(str1, str2):
    matcher = difflib.SequenceMatcher(None, str1, str2)
    return matcher.ratio() * 100

def compare_excel_files(file1, file2, output_file):
    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)
    
    ws1 = wb1.active
    ws2 = wb2.active
    
    wb3 = Workbook()
    ws3 = wb3.active
    ws3.append(["Name1", "Name2", "Similarity"])

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    unique_name = set()

    for row2 in ws2.iter_rows(min_row=2, min_col=3, max_row=ws2.max_row, max_col=3):
        cell2 = row2[0]
        name2 = cell2.value

        # Проверяем, есть ли уже name2 в unique_name
        if name2 in unique_name:
            continue

        unique_name.add(name2)
    
        for row1 in ws1.iter_rows(min_row=17, min_col=1, max_row=ws1.max_row, max_col=1):
            cell1 = row1[0]
            name1 = cell1.value
            
            similarity = similarity_percentage(name1, name2)
            
            if similarity > 85:
                new_row = [name1, name2, similarity]
                ws3.append(new_row)

                # Для дальнейшей оценки
                if similarity < 95:
                    # # Определяем координаты добавленной ячейки
                    # added_cell = ws3.cell(row=ws3.max_row, column=ws3.max_column)
                    # # Устанавливаем для добавленной ячейки желтый цвет
                    # added_cell.fill = yellow_fill

                    for cell in ws3[ws3.max_row]:
                        cell.fill = yellow_fill       

                print(name1, name2, similarity)
    
    wb3.save(output_file)

compare_excel_files("drug_side_effects_8.xlsx", "Частотный_словарь_побочных_эффектов.xlsx", "output.xlsx")
