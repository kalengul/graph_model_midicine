from openpyxl import Workbook
import os
from parse_html import drug

from googletrans import Translator
import time

translator = Translator()

def translate_text(text):
    while True:
        try:
            translation = translator.translate(text, dest='en')
            return translation.text
        except Exception as e:
            print(f"Ошибка перевода: {e}")
            time.sleep(0.5)  # Ждем перед повторной попыткой


# Получаем список файлов в папке
files = os.listdir('./drugs_html')

drug_list = []
for file_name in files:
    drug_list.append(drug(file_name))

# print(drug_list[0].file_name)
# print(drug_list[0].name_ru)
# print(drug_list[0].name_en)
# print(drug_list[0].number)

# Создание новой книги Excel
wb = Workbook()
sheet = wb.active

sheet.cell(row=1, column=1, value="Имя файла")  # Записываем названия классов в первую строку
sheet.cell(row=2, column=1, value="Название ru")  # Записываем названия классов в первую строку
sheet.cell(row=3, column=1, value="Название en")  # Записываем названия классов в первую строку
sheet.cell(row=4, column=1, value="Количество обращений")  # Записываем названия классов в первую строку

counter = 0

# Запись данных из массива классов в столбцы
# max_length = max(drug.get_len_list() for drug in drug_list)  # Находим максимальную длину массива данных
for i, drg in enumerate(drug_list):
    sheet.cell(row=1, column=i+2, value=drg.file_name)  # Записываем названия классов в первую строку
    sheet.cell(row=2, column=i+2, value=drg.name_ru)  # Записываем названия классов в первую строку
    sheet.cell(row=3, column=i+2, value=drg.name_en)  # Записываем названия классов в первую строку
    sheet.cell(row=4, column=i+2, value=drg.number)  # Записываем названия классов в первую строку
    offset = 5
    # Ваш код для записи в таблицу
    for j, drug_side_effect in enumerate(drg.drug_side_effect_list):

        # Перевод
        translated_name = translate_text(drug_side_effect.name)
        counter+= 1
        print("counter:", counter)

        value = f"{translated_name} ({drug_side_effect.percents}% , {drug_side_effect.cases})"
        sheet.cell(row=offset, column=i+2, value=value)  # Записываем данные класса в соответствующий столбец
        offset += 1 
        for k, side_effect in enumerate(drug_side_effect.sub_side_effect):

            # Перевод
            translated_sub_name = translate_text(side_effect["name"])
            counter+= 1
            print("counter:", counter)

            sub_value = f"{translated_sub_name} ({side_effect['cases']})"
            sheet.cell(row=offset, column=i+2, value=sub_value)  # Записываем данные класса в соответствующий столбец
            offset += 1

# Сохраняем книгу
wb.save('test.xlsx')